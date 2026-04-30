import openpyxl
import os

FILE_NAME = "Faltas.xlsx"

def init_excel():
    if not os.path.exists(FILE_NAME):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Faltas"
        ws.append(["Data", "Turma", "Nome"])
        wb.save(FILE_NAME)

def add_record(data, turma, nome):
    init_excel()
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active
        ws.append([data, turma, nome])
        wb.save(FILE_NAME)
        return True
    except Exception as e:
        print(f"Error adding record: {e}")
        return False

def add_multiple_records(records):
    init_excel()
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active
        for r in records:
            ws.append([r['data'], r['turma'], r['nome']])
        wb.save(FILE_NAME)
        return True
    except Exception as e:
        print(f"Error adding records: {e}")
        return False

def delete_record(data, turma, nome):
    init_excel()
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active
        found = False
        # Percorre de trás pra frente para evitar problema de índices ao deletar
        for row_idx in range(ws.max_row, 1, -1):
            c_data = ws.cell(row=row_idx, column=1).value
            c_turma = ws.cell(row=row_idx, column=2).value
            c_nome = ws.cell(row=row_idx, column=3).value
            
            if str(c_data) == str(data) and str(c_turma) == str(turma) and str(c_nome).strip().lower() == str(nome).strip().lower():
                ws.delete_rows(row_idx, 1)
                found = True
                break # Remove apenas a 1a ocorrência
                
        if found:
            wb.save(FILE_NAME)
            return True
        return False
    except Exception as e:
        print(f"Error deleting record: {e}")
        return False

def search_by_date(search_data):
    init_excel()
    results = []
    try:
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(search_data):
                results.append({
                    "Data": row[0],
                    "Turma": row[1],
                    "Nome": row[2]
                })
        return results
    except Exception as e:
        print(f"Error searching record: {e}")
        return results
